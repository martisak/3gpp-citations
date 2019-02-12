# -*- coding: utf-8 -*-

"""
This project aims to generate BiBTeX (http://www.bibtex.org/) files that
can be used when citing 3GPP (3gpp.org) specifications. The input is a document
list exported from the  3GPP Portal (https://portal.3gpp.org/).
"""

from __future__ import print_function
import argparse
from argparse import RawTextHelpFormatter
from datetime import datetime
from openpyxl import load_workbook
from bibtexparser.bwriter import BibTexWriter
from bibtexparser.bibdatabase import BibDatabase
from lxml import html
import requests
from tqdm import tqdm


DESCRIPTION = """
3GPP Bibtex entry generator --- Convert 3GPP document list from .xls to .bib.

1. Go to the [3GPP Portal](https://portal.3gpp.org/#55936-specifications)
2. Generate the list of specifications you want.
3. Download to Excel and save file
4. Run
   `python 3gpp-citations.py -i exported.xlsx -o 3gpp.bib`
5. Use in LaTeX.

* The output `bibtex` class is set to `@techreport`.
* The version and date are read from the URL, but it is slow so it takes
  a while to parse the list. If you find an easy solution to this, let me know.
"""

EPILOG = """

Example output:

@Techreport{3gpp.36.331,
  author = "3GPP",
  title = "{Evolved Universal Terrestrial Radio Access (E-UTRA);
        Radio Resource Control (RRC); Protocol specification}",
  type = "TS",
  institution = "{3rd Generation Partnership Project (3GPP)}",
  number = "{36.331}",
  days = 11,
  month = jul,
  year = 2016,
  url = "http://www.3gpp.org/dynareport/36331.htm",
}
"""


def parse_excel_row(row):
    """
    Parse a row in the sheet and return the data.
    """

    number = row[0].value
    title = row[2].value
    doctype = row[1].value

    return number, title, doctype


def format_entry(number, title, doctype, url):
    """
    Format the bibtex entry, return as dict
    """

    return {
        'ID': "3gpp.{}".format(number),
        'ENTRYTYPE': "techreport",
        'title': "{{{}}}".format(title),
        'type': doctype,
        'author': "3GPP",
        'institution': "{3rd Generation Partnership Project (3GPP)}",
        'number': number,
        'url': url
    }


def format_url(number, xelatex=True):
    r"""
    This function formats the URL field. If xelatex is used
    then we can allow for break-markers "\-"
    """

    # Disable Anomalous backslash in string: '\-'.
    # String constant might be missing an r prefix.
    # (anomalous-backslash-in-string)
    breakchar = "\-" if xelatex else ""  # pylint: disable=W1401

    url = "http://www.3gpp.org/{breakchar}DynaReport/" \
        "{breakchar}{number}.htm".format(
            breakchar=breakchar,
            number=number.replace(".", ""))

    return url


def parse_date(datestr):
    """
    This function parses a string of the form 1982-06-22 into
    year, month, day and returns them as strings.
    """

    datetime_object = datetime.strptime(datestr, '%Y-%m-%d')

    year = str(datetime_object.year)
    month = str(datetime_object.month)
    day = str(datetime_object.day)

    return year, month, day


def get_bibdatabase():
    """
    Create an empty BibDatabase
    """

    bib_database = BibDatabase()
    bib_database.entries = []

    return bib_database


def get_worksheet(filename):
    """
    Open a workbook and return the first sheet.
    """

    workbook = load_workbook(filename)
    worksheet = workbook[workbook.sheetnames[0]]

    return worksheet


def get_entry(row, xelatex=True):
    """
    Return an entry from a row in the Excel-sheet
    """

    number, title, doctype = parse_excel_row(row)

    if number is None:
        return None

    url = format_url(number, xelatex)
    entry = format_entry(number, title, doctype, url)

    # The Excel sheet does not contain version or release.
    if row[0].hyperlink is not None:
        # entry['url'] = row[0].hyperlink.target

        page = requests.get(row[0].hyperlink.target)
        tree = html.fromstring(page.content)

        # It is enough to go through the first two (latest two) releases.
        for release in range(2):

            release_row = tree.xpath(
                ('//tr[@id="SpecificationReleaseControl1_rpbReleases_i{}'
                 '_ctl00_specificationsVersionGrid_ctl00__0"]/td/div/a')
                .format(release))

            if release_row:
                daterow = tree.xpath(
                    ('//tr[@id="SpecificationReleaseControl1_rpbReleases'
                     '_i{}_ctl00_specificationsVersionGrid_ctl00__0"]/td')
                    .format(release))

                entry['note'] = "Version {}".format(
                    release_row[1].text.strip())

                datestr = daterow[2].text.strip()
                if datestr:
                    entry['year'], entry['month'], entry['day'] = \
                        parse_date(datestr)

                break
    return entry


def write_bibtex(bib_database, filename=None):
    """
    If a filename is submitted we print to file, otherwise to stdout
    """

    writer = BibTexWriter()

    if filename is not None:
        with open(filename, 'w') as bibfile:
            bibfile.write(writer.write(bib_database))
    else:
        print(writer.write(bib_database))


def main(args):
    """
    The main function that does all the heavy lifting.
    """

    bib_database = get_bibdatabase()
    worksheet = get_worksheet(args.input)

    # Iterate over the rows in the Excel-sheet but skip the header.
    for row in tqdm(
            worksheet.iter_rows(row_offset=1),
            total=worksheet.max_row - 1):

        entry = get_entry(row, args.xelatex)

        if entry is not None:
            bib_database.entries.append(entry)

    write_bibtex(bib_database, args.output)


def parse_args(args):
    """
    Parse arguments
    """

    parser = argparse.ArgumentParser(
        description=DESCRIPTION,
        epilog=EPILOG,
        formatter_class=RawTextHelpFormatter)

    parser.add_argument('--input', '-i', metavar='INPUT',
                        required=True,
                        help=('The Excel file generated by and '
                              'exported from the 3GPP Portal '
                              '(https://portal.3gpp.org)'))
    parser.add_argument('--output', '-o', metavar='OUTPUT',
                        help=('The bib file to write to. '
                              'STDOUT is used if omitted.'))
    parser.add_argument('--xelatex',
                        action='store_true',
                        help='Use line breaks')

    args = parser.parse_args(args)

    return args
