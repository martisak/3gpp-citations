from openpyxl import load_workbook
from bibtexparser.bwriter import BibTexWriter
from bibtexparser.bibdatabase import BibDatabase
import time
import sys
from lxml import html
import requests


# @Techreport{3gpp.36.331,
#   author = "3GPP",
#   title = "{Evolved Universal Terrestrial Radio Access (E-UTRA); Radio Resource Control (RRC); Protocol specification}",
#   type = "TS",
#   institution = "{3rd Generation Partnership Project (3GPP)}",
#   number = "{36.331}",
#   days = 11,
#   month = jul,
#   year = 2016,
#   url = "http://www.3gpp.org/dynareport/36331.htm",
# }

db = BibDatabase()
db.entries = []

wb2 = load_workbook(sys.argv[1])
sheetnames = wb2.get_sheet_names()
ws = wb2[sheetnames[0]]

# Iterate over the rows in the Excel-sheet but skip the header.
for row in ws.iter_rows(row_offset=1):

    number = row[0].value
    title = row[2].value
    type = row[1].value

    if number is None:
        continue

    entry = {
        'ID': "3gpp.{}".format(number),
        'ENTRYTYPE': "techreport",
        'title': "{{{}}}".format(title),
        'type': type,
        'author': "3GPP",
        'institution': "{3rd Generation Partnership Project (3GPP)}",
        'number': number
    }

    if row[0].hyperlink is not None:
        entry['url'] = row[0].hyperlink.target
        page = requests.get(entry['url'])
        tree = html.fromstring(page.content)

        for release in range(2):

            row = tree.xpath(
                '//tr[@id="SpecificationReleaseControl1_rpbReleases_i{}_ctl00_specificationsVersionGrid_ctl00__0"]/td/div/a'
                .format(release))

            if len(row) > 0:
                daterow = tree.xpath(
                    '//tr[@id="SpecificationReleaseControl1_rpbReleases_i{}_ctl00_specificationsVersionGrid_ctl00__0"]/td'
                    .format(release))

                entry['note'] = "Version {}".format(row[1].text.strip())

                if daterow[2].text.strip() is not "":
                    date = daterow[2].text.split('-')
                    entry['day'] = date[2].strip()
                    entry['year'] = date[0].strip()
                    entry['month'] = date[1].strip()
                break

    print entry
    db.entries.append(entry)

writer = BibTexWriter()
with open(sys.argv[2], 'w') as bibfile:
    bibfile.write(writer.write(db))
